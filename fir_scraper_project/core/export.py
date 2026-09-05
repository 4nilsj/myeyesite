"""Export utilities for generating CSV and formatted Excel (.xlsx) spreadsheets."""

import csv
import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.classifier import get_crime_type_badge
from core.config import _extract_fir_number_from_filename

HEADERS = [
    "FIR No.",
    "Filename",
    "Station ID",
    "Police Station",
    "Complaint Date",
    "Crime Category",
    "Acts & Sections",
    "Incident Location",
    "Taluk (TQ)",
    "Complainant Name",
    "Complainant Address",
    "Accused Name",
    "Accused Address",
    "Victim Name",
    "Victim Address",
    "Pages",
    "Size (MB)",
    "OCR Used",
    "Case Summary",
]


def _build_row_data(record: dict[str, Any]) -> list[Any]:
    filename = record.get("name", "")
    fir_num = _extract_fir_number_from_filename(filename)
    fir_str = f"#{fir_num:04d}" if fir_num is not None else "—"

    acts = record.get("acts_sections", "")
    text = record.get("text", "")
    crime_type = get_crime_type_badge(acts, text)

    summary = record.get("plain_summary") or record.get("summary") or ""
    summary_clean = " ".join(summary.splitlines()).strip()

    return [
        fir_str,
        filename,
        record.get("station_id", ""),
        record.get("station_name", ""),
        record.get("complaint_date") or record.get("parsed_date") or "—",
        crime_type,
        acts or "Not specified",
        record.get("location", ""),
        record.get("tq", ""),
        record.get("complainant_name", ""),
        record.get("complainant_address", ""),
        record.get("accused_name", ""),
        record.get("accused_address", ""),
        record.get("victim_name", ""),
        record.get("victim_address", ""),
        record.get("pages", 1),
        record.get("size_mb", 0.0),
        "Yes" if record.get("ocr_used") else "No",
        summary_clean,
    ]


def generate_csv_data(records: list[dict[str, Any]]) -> str:
    """Generate UTF-8 CSV content with Byte Order Mark (BOM) for Excel compatibility."""
    output = io.StringIO()
    # Write UTF-8 BOM so Excel opens Kannada and Unicode characters properly
    output.write("\ufeff")
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(HEADERS)

    for rec in records:
        writer.writerow(_build_row_data(rec))

    return output.getvalue()


def generate_excel_workbook(records: list[dict[str, Any]]) -> bytes:
    """Generate a professionally styled Excel (.xlsx) workbook as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FIR Crime Records"

    # Header styling (Deep Navy with White text)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Borders
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # Alternate row fill
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    regular_font = Font(name="Calibri", size=10)
    top_align = Alignment(vertical="top", wrap_text=True)

    # Write headers
    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    # Write data rows
    for row_idx, rec in enumerate(records, start=2):
        row_values = _build_row_data(rec)
        ws.append(row_values)
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.alignment = top_align
            cell.border = thin_border
            if is_even:
                cell.fill = alt_fill

    # Auto-adjust column widths
    max_widths = {
        1: 10,   # FIR No
        2: 24,   # Filename
        3: 12,   # Station ID
        4: 26,   # Station Name
        5: 16,   # Date
        6: 22,   # Crime Type
        7: 28,   # Acts
        8: 24,   # Location
        9: 14,   # TQ
        10: 22,  # Complainant
        11: 30,  # Comp Address
        12: 22,  # Accused
        13: 30,  # Accused Address
        14: 20,  # Victim
        15: 28,  # Victim Address
        16: 10,  # Pages
        17: 12,  # Size MB
        18: 12,  # OCR
        19: 45,  # Summary
    }

    for col_idx, width in max_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
