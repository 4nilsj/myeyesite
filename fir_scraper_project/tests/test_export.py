import io
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

# Ensure root project directory is in sys.path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import fitz
import openpyxl

import app
from core.db import _init_db, _save_record_to_db
from core.export import generate_csv_data, generate_excel_workbook


class TestExportAndDossier(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.addCleanup(lambda: shutil.rmtree(self.temp_dir, ignore_errors=True))

        self.test_db_path = Path(self.temp_dir) / "test_export.db"
        self.test_pdf_dir = Path(self.temp_dir) / "pdfs"
        self.test_pdf_dir.mkdir(parents=True, exist_ok=True)

        self.original_db_file = app.DB_FILE
        self.original_pdf_dir = app.PDF_DIR

        app.DB_FILE = self.test_db_path
        app.PDF_DIR = self.test_pdf_dir

        _init_db()

        self.sample_record = {
            "name": "fir_ps717_0001.pdf",
            "station_id": "717",
            "station_name": "Madbool Station (717)",
            "complaint_date": "10/05/2024",
            "parsed_date": "2024-05-10",
            "location": "Madbool Main Road",
            "tq": "chittapur",
            "complainant_name": "Ramesh Kumar",
            "complainant_address": "Madbool Village",
            "accused_name": "Suresh Gowda",
            "accused_address": "Kalagi",
            "victim_name": "Ramesh Kumar",
            "victim_address": "Madbool Village",
            "pages": 2,
            "size_mb": 0.35,
            "ocr_used": 0,
            "ocr_status": "Direct",
            "summary": "Theft of motorcycle from front yard",
            "plain_summary": "Theft of motorcycle from front yard",
            "acts_sections": "303 BNS",
            "text": "Full complaint text for motorcycle theft",
        }
        # Create physical dummy PDF so single dossier / pdf detail tests find it
        dummy_pdf_path = self.test_pdf_dir / "fir_ps717_0001.pdf"
        doc = fitz.open()
        page = doc.new_page()
        page.insert_text((50, 50), "Dummy FIR PDF Content")
        doc.save(str(dummy_pdf_path))
        doc.close()

        stat = dummy_pdf_path.stat()
        _save_record_to_db(self.sample_record, mtime=int(stat.st_mtime), size=stat.st_size)

        app.app.config["TESTING"] = True
        self.client = app.app.test_client()

    def tearDown(self):
        app.DB_FILE = self.original_db_file
        app.PDF_DIR = self.original_pdf_dir

    def test_generate_csv_data(self):
        csv_str = generate_csv_data([self.sample_record])
        # Check UTF-8 BOM
        self.assertTrue(csv_str.startswith("\ufeff"))
        self.assertIn("FIR No.", csv_str)
        self.assertIn("fir_ps717_0001.pdf", csv_str)
        self.assertIn("Ramesh Kumar", csv_str)
        self.assertIn("Theft", csv_str)

    def test_generate_excel_workbook(self):
        xlsx_bytes = generate_excel_workbook([self.sample_record])
        self.assertIsInstance(xlsx_bytes, bytes)
        self.assertGreater(len(xlsx_bytes), 1000)

        # Verify workbook is readable by openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        self.assertIn("FIR Crime Records", wb.sheetnames)
        ws = wb["FIR Crime Records"]
        self.assertEqual(ws.cell(row=1, column=1).value, "FIR No.")
        self.assertEqual(ws.cell(row=2, column=2).value, "fir_ps717_0001.pdf")
        self.assertEqual(ws.cell(row=2, column=6).value, "Theft")

    def test_export_csv_endpoint(self):
        response = self.client.get("/export/csv?station_id=717")
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response.headers.get("Content-Type", ""))
        self.assertIn("attachment; filename=", response.headers.get("Content-Disposition", ""))
        self.assertIn(b"fir_ps717_0001.pdf", response.data)

    def test_export_excel_endpoint(self):
        response = self.client.get("/export/excel?station_id=717")
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response.headers.get("Content-Type", ""))
        self.assertIn("attachment; filename=", response.headers.get("Content-Disposition", ""))
        self.assertGreater(len(response.data), 1000)

    def test_batch_dossier_endpoint(self):
        response = self.client.get("/dossier?station_id=717")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Crime Intelligence Digest", response.data)
        self.assertIn(b"fir_ps717_0001.pdf", response.data)

    def test_single_dossier_endpoint(self):
        response = self.client.get("/dossier/fir_ps717_0001.pdf")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Official Police Case Dossier", response.data)
        self.assertIn(b"Ramesh Kumar", response.data)
        self.assertIn(b"Investigating Officer", response.data)

    def test_extract_rich_dossier_data(self):
        from core.dossier_extractor import extract_rich_dossier_data

        sample = {
            "complainant_name": "Liyakat Ahmad",
            "complainant_address": "Pethshiroor",
            "accused_name": "Unknown",
            "victim_name": "Not clearly found",
            "location": "Kalaburgi Sedam Road",
            "tq": "kalagi",
            "text": (
                "3. (a) ಕೃತ್ಯ ನಡೆದ ದಿನ : Sunday ದಿನಾಂಕ ದಿಂದ : 06/06/2026 ದಿನಾಂಕ ವರೆಗೆ : 06/06/2026 ವೇಳೆಯಿಂದ : 21:00:00 ವೇಳೆಯವರೆಗೆ : 21:05:00\n"
                "(b) ಠಾಣೆಯಲ್ಲಿ ವರ್ತಮಾನ ಸ್ವೀಕರಿಸಿದ ದಿನಾಂಕ : 07/06/2026 ಬರವಣಿಗೆಯಲ್ಲಿ/ಹೇಳಿಕೆ : Written 15:00:00\n"
                "4. (b) ಪೊಲೀಸ್‌ಠಾಣೆ ಯಿಂದ ಇರುವ ದಿಕ್ಕು ಮತ್ತು ದೂರ : Towards East 4.000 Km Madbool PS\n"
                "5. ಪಿರ್ಯಾದುದಾರ / ಬಾತ್ಮೀದಾರ : (a) ಹೆಸರು : Liyakat ahmad ತಂದೆ/ಗಂಡನ ಹೆಸರು : Rajaq patel (b) ವಯಸ್ಸು : 26 (g) ದೂರವಾಣಿ : 7795441438 (h) ಲಿಂಗ : Male\n"
                "6. ಗೊತ್ತಿರುವ/ಅನುಮಾನಿತ /ಅಪರಿಚಿತ ವ್ಯಕ್ತಿಯ ಪೂರ್ತಿ ವಿವರಗಳು : ಹೆಸರು / ಸ.ನಂ 1 Driver of the XUV Car no KA32MA5210 (A1) Accused Adult Male\n"
                "7. ನೊಂದವರ ವಿವರಗಳು : ಸ.ನಂ ಹೆಸರು 1 Smt Ayesha siddiq Grievous Female 23 Housewife\n"
                "8. ಕಳುವಾಗಿರುವ / ಬಾಗಿಯಾಗಿರುವ ಸ್ವತ್ತುಗಳ ವಿವರಗಳು : 1 Other Property Lorry No KA32AA3076 300000.00 ಕಳುವಾಗಿರುವ / ಬಾಗಿಯಾಗಿರುವ ಸ್ವತ್ತುಗಳ ಮೌಲ್ಯ : 300000.00\n"
                "10. ಪ್ರಥಮ ವರ್ತಮಾನ ವರದಿಯ ವಿವರಗಳು : Liyakat Ahmad reported that while traveling on Kalaburgi-Sedam road near Madbool cross an XUV car collided from behind.\n"
                "11. ಕ್ರಮ ತೆಗೆದುಕೊಂಡ ಬಗ್ಗೆ ವಿವರ :"
            ),
        }
        res = extract_rich_dossier_data(sample)
        self.assertEqual(res["complainant"]["father_spouse"], "Rajaq patel")
        self.assertEqual(res["complainant"]["age"], "26")
        self.assertEqual(res["complainant"]["phone"], "7795441438")
        self.assertIn("Sunday,", res["incident_meta"]["occurrence_datetime"])
        self.assertIn("07/06/2026 at 15:00:00", res["incident_meta"]["ps_received_datetime"])
        self.assertIn("Driver of the XUV Car", res["accused_list"][0]["name"])
        self.assertIn("Smt Ayesha siddiq", res["victim_list"][0]["name"])
        self.assertEqual(len(res["property_items"]), 1)
        self.assertIn("Kalaburgi-Sedam", res["fir_narrative"])


if __name__ == "__main__":
    unittest.main()
